"""Unit tests for xlsx_gen."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.tools.builtin.docgen.xlsx_gen import gen


@pytest.mark.asyncio
async def test_xlsx_gen_dimensions(tmp_path: Path):
    out = tmp_path / "data.xlsx"
    sheet = {
        "headers": ["A", "B", "C"],
        "rows": [[1, 2, 3], [4, 5, 6], [7, 8, 9], [10, 11, 12], [13, 14, 15]],
        "sheet_name": "Metrics",
    }
    path = await gen(sheet, str(out))
    wb = load_workbook(path)
    ws = wb["Metrics"]
    # header + 5 rows, 3 cols
    assert ws.max_row == 6
    assert ws.max_column == 3
    assert [c.value for c in ws[1]] == ["A", "B", "C"]
